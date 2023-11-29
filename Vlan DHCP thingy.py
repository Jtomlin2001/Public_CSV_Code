import meraki
import requests
import json
from openpyxl import Workbook

wb = Workbook()

# grab the active worksheet
ws = wb.active


base_url = 'https://api.meraki.com/api/v1'
session_params = {
    'api_key': api_key,
    'base_url': base_url,

}
dashboard = meraki.DashboardAPI(**session_params)
headers = {
        'X-Cisco-Meraki-API-Key': api_key
    }
network_devices = []
device = []   
organization_id = '445912'
networks = dashboard.organizations.getOrganizationNetworks(organization_id)
time = 0
test = 0

for network in networks:
    ws = wb.create_sheet(f"{network['name']}")
    network_id = network['id']
    url =  requests.get(f"{base_url}/networks/{network_id}/appliance/vlans", headers={'Authorization': f'Bearer {api_key}'})
    vlan_data = (url.json())
    vlan_count = 0
    for vlan in vlan_data:
        try:
            print(str(vlan['dhcpRelayServerIps']))
        except:
            print("test")
        try:
            vlan_id = str({vlan['id']}).replace("{","").replace("}", "")
            print(vlan_id)
            vlan_count += 1
            vlan_name = str({vlan['name']}).replace("{","").replace("}", "")
            vlan_dhcp_handling = str({vlan['dhcpHandling']}).replace("{","").replace("}", "")
            vlan_subnet = str({vlan['subnet']}).replace("{","").replace("}", "")
            ws['A1'] = "Vlan Name"
            ws['B1'] = "Vlan ID"
            ws['C1'] = "DHCP Handling"
            ws['D1'] = "DHCP Relay Server IPs"
            ws.cell(row=1+vlan_count, column=1, value = vlan_id)
            ws.cell(row=1+vlan_count, column=2, value = vlan_name)
            ws.cell(row=1+vlan_count, column=3, value = vlan_dhcp_handling)
            ws.cell(row=1+vlan_count, column=4, value = "")
            try:
                vlan_dhcp_relay = (str(vlan['dhcpRelayServerIps']))
                relay_ip = vlan_dhcp_relay.replace("[","").replace("]", "").replace("'","")
                json = {
                    "dhcpRelayServerIps": [
                    relay_ip,
                    "172.29.16.71"
                    ]
                }
                print(json)
                ######################PUT REQUEST ########################################
                url2 =  requests.put(f"{base_url}/networks/{network_id}/appliance/vlans/{vlan_id}", headers={'Authorization': f'Bearer {api_key}'}, json=json)
                print(f"{network['name']}")
                if (url2.status_code) == 200:
                    print(f"{network['name']} dhcp server ips updated to {relay_ip} and 172.29.16.71")
                else:
                    print(f"{network['name']}")
                ######################PUT REQUEST ########################################
            except:
                print()
            wb.save('Vlan_Info.xlsx')
        except:
            print(f"{network['name']} Has no Vlans")
#    network_devices.append(url.json())
#    time +=1
#    print(time)
#for device in network_devices:
#    for info in device:
#        model = info['model']
#        if model.startswith("MX"): 
#            print(info['name'])
#            serial = (info['serial'])
#            print(serial)
#            subnets = requests.get(f"https://api.meraki.com/api/v1/devices/{serial}/appliance/dhcp/subnets", headers={'Authorization': f'Bearer {api_key}'})
#            print(subnets.json())
#                
#
#
#
###vlans = Merakiclient.vlans.getNetworkVlans(network['id'])
#    for vlan in vlans:
#        print('Processing Vlan: {} - {}'.format(vlan['id'], vlan['name']))
#        if vlan['dhcpHandling'] == 'Relay DHCP to another server':
